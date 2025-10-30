from datetime import datetime

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.db.models import Avg, Count, ExpressionWrapper, F, FloatField, Max, Q, Value
from django.db.models import Prefetch
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    FormView,
    ListView,
    TemplateView,
    UpdateView,
)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ml_models.potential_evaluator import evaluate_player_potential
from performance.forms import AthleteForm
from performance.models import Athlete
from scouting.forms import ScoutedPlayerForm, ScoutingReportForm
from scouting.models import ScoutedPlayer, ScoutingReport


class ScoutedPlayerListView(LoginRequiredMixin, ListView):
    """Display list of scouted players with search and filters."""

    model = ScoutedPlayer
    template_name = 'scouting/scoutedplayer_list.html'
    context_object_name = 'players'
    paginate_by = 20
    login_url = reverse_lazy('accounts:login')

    def get_queryset(self):
        queryset = (
            super()
            .get_queryset()
            .select_related('created_by')
        )

        search_query = self.request.GET.get('q', '').strip()
        if search_query:
            queryset = queryset.filter(name__icontains=search_query)

        club_query = self.request.GET.get('club', '').strip()
        if club_query:
            queryset = queryset.filter(current_club__icontains=club_query)

        position_filter = self.request.GET.get('position', '').strip()
        if position_filter:
            queryset = queryset.filter(position=position_filter)

        status_filter = self.request.GET.get('status', '').strip()
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        order_by = self.request.GET.get('order', '').strip()
        allowed_ordering = {
            'market_value',
            '-market_value',
            'name',
            '-name',
            'birth_date',
            '-birth_date',
            'created_at',
            '-created_at',
        }
        if order_by in allowed_ordering:
            queryset = queryset.order_by(order_by)
        else:
            queryset = queryset.order_by('-created_at')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(
            {
                'position_choices': ScoutedPlayer.POSITION_CHOICES,
                'status_choices': ScoutedPlayer.STATUS_CHOICES,
                'order_choices': ScoutedPlayerForm.ORDER_BY_CHOICES,
                'current_search': self.request.GET.get('q', ''),
                'current_club': self.request.GET.get('club', ''),
                'current_position': self.request.GET.get('position', ''),
                'current_status': self.request.GET.get('status', ''),
                'current_order': self.request.GET.get('order', ''),
            }
        )
        return context


class PlayerComparisonView(LoginRequiredMixin, TemplateView):
    """Compare up to three scouted players side by side."""

    template_name = 'scouting/player_comparison.html'
    login_url = reverse_lazy('accounts:login')
    score_labels_map = {
        'technical_score': 'Técnica',
        'physical_score': 'Física',
        'tactical_score': 'Tática',
        'mental_score': 'Mental',
        'potential_score': 'Potencial',
    }

    def _extract_player_ids(self):
        """Return sanitized list of selected player IDs preserving order."""
        raw_ids = self.request.GET.getlist('players')
        sanitized = []
        for raw_id in raw_ids:
            if raw_id.isdigit():
                sanitized.append(int(raw_id))
        return sanitized

    def get(self, request, *args, **kwargs):
        player_ids = self._extract_player_ids()
        if not player_ids:
            messages.warning(request, 'Selecione ao menos dois jogadores para comparar.')
            return redirect('scouting:player_list')
        if len(player_ids) < 2:
            messages.warning(request, 'Selecione ao menos dois jogadores para comparar.')
            return redirect('scouting:player_list')
        if len(player_ids) > 3:
            messages.warning(
                request,
                'É possível comparar no máximo 3 jogadores. Os três primeiros selecionados foram utilizados.',
            )
            player_ids = player_ids[:3]
        self.selected_ids = player_ids
        return super().get(request, *args, **kwargs)

    def get_players(self):
        """Fetch players including their latest reports keeping selection order."""
        reports_prefetch = Prefetch(
            'reports',
            queryset=ScoutingReport.objects.select_related('created_by').order_by('-report_date', '-created_at'),
            to_attr='ordered_reports',
        )
        players_qs = (
            ScoutedPlayer.objects.filter(pk__in=self.selected_ids)
            .select_related('created_by')
            .prefetch_related(reports_prefetch)
        )
        players = list(players_qs)
        if len(players) != len(self.selected_ids):
            messages.warning(
                self.request,
                'Alguns jogadores selecionados não foram encontrados. Compare novamente se necessário.',
            )
        players.sort(key=lambda player: self.selected_ids.index(player.pk))
        if len(players) < 2:
            messages.warning(
                self.request,
                'Não foi possível carregar jogadores suficientes para a comparação.',
            )
            return []
        return players

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        score_fields = ScoutingReportForm.SCORE_FIELDS
        context.update(
            {
                'comparison_entries': [],
                'score_rows': [],
                'score_labels': [self.score_labels_map[field] for field in score_fields],
                'score_fields': score_fields,
                'chart_datasets': [],
                'has_reports': False,
            }
        )
        players = self.get_players()
        if len(players) < 2:
            return context

        comparison_entries = []
        for player in players:
            reports = getattr(player, 'ordered_reports', [])
            last_report = reports[0] if reports else None
            scores = {
                field: getattr(last_report, field, None) if last_report else None
                for field in score_fields
            }
            comparison_entries.append(
                {
                    'player': player,
                    'last_report': last_report,
                    'scores': scores,
                }
            )

        score_rows = []
        for field in score_fields:
            values = []
            best_value = None
            for entry in comparison_entries:
                value = entry['scores'].get(field)
                values.append(value)
                if value is not None:
                    best_value = value if best_value is None else max(best_value, value)
            score_rows.append(
                {
                    'field': field,
                    'label': self.score_labels_map[field],
                    'values': values,
                    'best_value': best_value,
                }
            )

        chart_palette = [
            {
                'border': 'rgba(16, 185, 129, 0.8)',
                'background': 'rgba(16, 185, 129, 0.25)',
                'point': '#10b981',
            },
            {
                'border': 'rgba(59, 130, 246, 0.8)',
                'background': 'rgba(59, 130, 246, 0.25)',
                'point': '#3b82f6',
            },
            {
                'border': 'rgba(236, 72, 153, 0.8)',
                'background': 'rgba(236, 72, 153, 0.25)',
                'point': '#ec4899',
            },
        ]
        chart_datasets = []
        for idx, entry in enumerate(comparison_entries):
            palette = chart_palette[idx % len(chart_palette)]
            data = [
                entry['scores'].get(field) if entry['scores'].get(field) is not None else 0
                for field in score_fields
            ]
            has_values = any(entry['scores'].get(field) is not None for field in score_fields)
            chart_datasets.append(
                {
                    'label': entry['player'].name,
                    'data': data,
                    'borderColor': palette['border'],
                    'backgroundColor': palette['background'],
                    'pointBackgroundColor': palette['point'],
                    'pointBorderColor': palette['border'],
                    'hidden': not has_values,
                }
            )

        context.update(
            {
                'comparison_entries': comparison_entries,
                'score_rows': score_rows,
                'chart_datasets': chart_datasets,
                'has_reports': any(
                    entry['last_report'] is not None for entry in comparison_entries
                ),
            }
        )
        return context


class ConvertToAthleteView(LoginRequiredMixin, SuccessMessageMixin, FormView):
    """Convert a scouted player into an athlete inside the performance module."""

    POSITION_MAP = {
        ScoutedPlayer.POSITION_GOALKEEPER: Athlete.POSITION_GOALKEEPER,
        ScoutedPlayer.POSITION_RIGHT_BACK: Athlete.POSITION_DEFENDER,
        ScoutedPlayer.POSITION_LEFT_BACK: Athlete.POSITION_DEFENDER,
        ScoutedPlayer.POSITION_CENTER_BACK: Athlete.POSITION_DEFENDER,
        ScoutedPlayer.POSITION_DEFENSIVE_MIDFIELDER: Athlete.POSITION_MIDFIELDER,
        ScoutedPlayer.POSITION_CENTRAL_MIDFIELDER: Athlete.POSITION_MIDFIELDER,
        ScoutedPlayer.POSITION_ATT_MIDFIELDER: Athlete.POSITION_MIDFIELDER,
        ScoutedPlayer.POSITION_RIGHT_WINGER: Athlete.POSITION_FORWARD,
        ScoutedPlayer.POSITION_LEFT_WINGER: Athlete.POSITION_FORWARD,
        ScoutedPlayer.POSITION_STRIKER: Athlete.POSITION_FORWARD,
    }
    POSITION_LABELS = dict(Athlete.POSITION_CHOICES)

    template_name = 'scouting/convert_to_athlete.html'
    form_class = AthleteForm
    login_url = reverse_lazy('accounts:login')
    success_message = 'Jogador contratado e integrado ao elenco com sucesso!'

    def dispatch(self, request, *args, **kwargs):
        self.scouted_player = get_object_or_404(ScoutedPlayer, pk=kwargs['pk'])
        if self.scouted_player.status == ScoutedPlayer.STATUS_HIRED:
            messages.info(
                request,
                'Este jogador já está marcado como contratado e disponível no módulo de performance.',
            )
            return redirect('scouting:player_detail', pk=self.scouted_player.pk)
        return super().dispatch(request, *args, **kwargs)

    def get_initial(self):
        initial = super().get_initial()
        mapped_position = self.POSITION_MAP.get(self.scouted_player.position)
        initial.update(
            {
                'name': self.scouted_player.name,
                'birth_date': self.scouted_player.birth_date,
                'nationality': self.scouted_player.nationality,
            }
        )
        if mapped_position:
            initial['position'] = mapped_position
        if self.scouted_player.market_value is not None:
            initial['market_value'] = self.scouted_player.market_value
        return initial

    def form_valid(self, form):
        athlete = form.save(commit=False)
        athlete.created_by = self.request.user
        if athlete.market_value is None and self.scouted_player.market_value is not None:
            athlete.market_value = self.scouted_player.market_value
        athlete.save()
        form.save_m2m()
        self.scouted_player.status = ScoutedPlayer.STATUS_HIRED
        if not self.scouted_player.hire_date:
            self.scouted_player.hire_date = timezone.localdate()
        self.scouted_player.save(update_fields=['status', 'hire_date', 'updated_at'])
        self.athlete = athlete
        messages.success(self.request, self.success_message)
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('performance:athlete_detail', args=[self.athlete.pk])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        mapped_position = self.POSITION_MAP.get(self.scouted_player.position)
        context.update(
            {
                'scouted_player': self.scouted_player,
                'suggested_position': mapped_position,
                'suggested_position_label': self.POSITION_LABELS.get(mapped_position),
            }
        )
        return context


class ScoutedPlayerCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    """Create a new scouted player record."""

    model = ScoutedPlayer
    form_class = ScoutedPlayerForm
    template_name = 'scouting/scoutedplayer_form.html'
    success_url = reverse_lazy('scouting:player_list')
    success_message = 'Jogador observado cadastrado com sucesso.'
    login_url = reverse_lazy('accounts:login')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class ScoutedPlayerDetailView(LoginRequiredMixin, DetailView):
    """Display details about a single scouted player."""

    model = ScoutedPlayer
    template_name = 'scouting/scoutedplayer_detail.html'
    context_object_name = 'player'
    login_url = reverse_lazy('accounts:login')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['has_photo'] = bool(self.object.photo)
        reports = (
            self.object.reports.select_related('created_by')
            .order_by('-report_date', '-created_at')
        )
        context['reports'] = reports
        score_fields = ScoutingReportForm.SCORE_FIELDS
        score_labels = {
            'technical_score': 'Técnica',
            'physical_score': 'Física',
            'tactical_score': 'Tática',
            'mental_score': 'Mental',
            'potential_score': 'Potencial',
        }
        context['score_labels'] = score_labels
        context['score_fields'] = score_fields
        reports_count = reports.count()
        context['reports_count'] = reports_count
        if reports_count:
            averages = reports.aggregate(
                **{f'{field}_avg': Avg(field) for field in score_fields}
            )
            average_values = [
                averages.get(f'{field}_avg') or 0 for field in score_fields
            ]
            overall_average = (
                sum(average_values) / len(score_fields) if reports_count else 0
            )
            score_averages = []
            for field in score_fields:
                avg_value = averages.get(f'{field}_avg') or 0
                score_averages.append(
                    {
                        'field': field,
                        'label': score_labels[field],
                        'average': avg_value,
                        'percentage': (avg_value / 10) * 100 if avg_value else 0,
                    }
                )
            context['score_averages'] = score_averages
            context['overall_average'] = overall_average
            timeline = list(
                reports.order_by('report_date', 'created_at')
            )
            context['timeline_labels'] = [
                report.report_date.strftime('%d/%m/%Y') for report in timeline
            ]
            context['timeline_values'] = [
                report.overall_score() for report in timeline
            ]
            context['show_timeline_chart'] = len(timeline) > 1
        else:
            context['score_averages'] = []
            context['overall_average'] = None
            context['timeline_labels'] = []
            context['timeline_values'] = []
            context['show_timeline_chart'] = False
        context['potential_prediction'] = evaluate_player_potential(self.object)
        context['can_convert'] = self.object.status != ScoutedPlayer.STATUS_HIRED
        context['convert_url'] = reverse('scouting:player_convert', args=[self.object.pk])
        return context


class ScoutedPlayerUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    """Edit an existing scouted player."""

    model = ScoutedPlayer
    form_class = ScoutedPlayerForm
    template_name = 'scouting/scoutedplayer_form.html'
    success_message = 'Jogador observado atualizado com sucesso.'
    login_url = reverse_lazy('accounts:login')

    def get_success_url(self):
        return reverse_lazy('scouting:player_detail', args=[self.object.pk])


class ScoutedPlayerDeleteView(LoginRequiredMixin, DeleteView):
    """Remove a scouted player after confirmation."""

    model = ScoutedPlayer
    template_name = 'scouting/scoutedplayer_confirm_delete.html'
    success_url = reverse_lazy('scouting:player_list')
    login_url = reverse_lazy('accounts:login')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Jogador observado removido com sucesso.')
        return super().delete(request, *args, **kwargs)


class ScoutingDashboardView(LoginRequiredMixin, TemplateView):
    """Display consolidated scouting metrics and insights."""

    template_name = 'scouting/scouting_dashboard.html'
    login_url = reverse_lazy('accounts:login')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        players = ScoutedPlayer.objects.select_related('created_by')
        reports = ScoutingReport.objects.select_related('player', 'created_by')

        total_players = players.count()
        total_reports = reports.count()
        user_model = get_user_model()
        active_scouts = (
            user_model.objects.filter(scouting_reports__isnull=False)
            .distinct()
            .count()
        )

        status_totals = {
            item['status']: item['total']
            for item in players.values('status').annotate(total=Count('id'))
        }
        status_distribution = []
        negotiating_count = status_totals.get(ScoutedPlayer.STATUS_NEGOTIATING, 0)
        for code, label in ScoutedPlayer.STATUS_CHOICES:
            count = status_totals.get(code, 0)
            percentage = (count / total_players * 100) if total_players else 0
            status_distribution.append(
                {
                    'code': code,
                    'label': label,
                    'count': count,
                    'percentage': percentage,
                }
            )

        pipeline_efficiency = (
            negotiating_count / total_players * 100 if total_players else 0
        )

        annotated_players = players.annotate(
            report_count=Count('reports', distinct=True),
            technical_avg=Avg('reports__technical_score'),
            physical_avg=Avg('reports__physical_score'),
            tactical_avg=Avg('reports__tactical_score'),
            mental_avg=Avg('reports__mental_score'),
            potential_avg=Avg('reports__potential_score'),
            last_report_date=Max('reports__report_date'),
        ).annotate(
            overall_avg=ExpressionWrapper(
                (
                    Coalesce(F('technical_avg'), Value(0.0))
                    + Coalesce(F('physical_avg'), Value(0.0))
                    + Coalesce(F('tactical_avg'), Value(0.0))
                    + Coalesce(F('mental_avg'), Value(0.0))
                    + Coalesce(F('potential_avg'), Value(0.0))
                )
                / Value(5.0),
                output_field=FloatField(),
            ),
            potential_avg_value=Coalesce(F('potential_avg'), Value(0.0)),
        ).annotate(
            overall_percentage=ExpressionWrapper(
                F('overall_avg') * Value(10),
                output_field=FloatField(),
            ),
        )

        top_players = (
            annotated_players.filter(report_count__gt=0)
            .order_by('-overall_avg', '-last_report_date')[:5]
        )

        priority_players = (
            annotated_players.filter(report_count__gt=0)
            .filter(
                Q(status=ScoutedPlayer.STATUS_MONITORING)
                | Q(status=ScoutedPlayer.STATUS_INTERESTED)
            )
            .order_by('-potential_avg_value', '-overall_avg')[:5]
        )

        latest_reports = reports.order_by('-report_date', '-created_at')[:6]

        context.update(
            {
                'total_players': total_players,
                'total_reports': total_reports,
                'active_scouts': active_scouts,
                'status_distribution': status_distribution,
                'status_labels': [item['label'] for item in status_distribution],
                'status_values': [item['count'] for item in status_distribution],
                'top_players': top_players,
                'priority_players': priority_players,
                'latest_reports': latest_reports,
                'has_reports': total_reports > 0,
                'pipeline_efficiency': pipeline_efficiency,
            }
        )
        return context


class ScoutingReportListView(LoginRequiredMixin, ListView):
    """Display all scouting reports with filtering options."""

    model = ScoutingReport
    template_name = 'scouting/scoutingreport_list.html'
    context_object_name = 'reports'
    paginate_by = 20
    login_url = reverse_lazy('accounts:login')

    def get_queryset(self):
        queryset = (
            super()
            .get_queryset()
            .select_related('player', 'created_by')
        )
        player_id = self.request.GET.get('player')
        if player_id:
            queryset = queryset.filter(player_id=player_id)

        scout_id = self.request.GET.get('scout')
        if scout_id:
            queryset = queryset.filter(created_by_id=scout_id)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_model = get_user_model()
        context.update(
            {
                'players': ScoutedPlayer.objects.order_by('name'),
                'scouts': user_model.objects.filter(
                    scouting_reports__isnull=False
                )
                .distinct()
                .order_by('first_name', 'last_name', 'email'),
                'current_player': self.request.GET.get('player', ''),
                'current_scout': self.request.GET.get('scout', ''),
            }
        )
        return context


class ScoutingReportCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    """Register a new scouting report."""

    model = ScoutingReport
    form_class = ScoutingReportForm
    template_name = 'scouting/scoutingreport_form.html'
    success_message = 'Relatório de scouting criado com sucesso.'
    login_url = reverse_lazy('accounts:login')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def get_initial(self):
        initial = super().get_initial()
        player_id = self.request.GET.get('player')
        if player_id:
            initial['player'] = player_id
        return initial

    def get_success_url(self):
        return reverse('scouting:report_detail', args=[self.object.pk])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['score_fields'] = ScoutingReportForm.SCORE_FIELDS
        return context


class ScoutingReportDetailView(LoginRequiredMixin, DetailView):
    """Show detailed information about a scouting report."""

    model = ScoutingReport
    template_name = 'scouting/scoutingreport_detail.html'
    context_object_name = 'report'
    login_url = reverse_lazy('accounts:login')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['score_labels'] = [
            'Técnica',
            'Física',
            'Tática',
            'Mental',
            'Potencial',
        ]
        context['score_values'] = [
            self.object.technical_score,
            self.object.physical_score,
            self.object.tactical_score,
            self.object.mental_score,
            self.object.potential_score,
        ]
        return context


class ScoutingReportUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    """Update an existing scouting report."""

    model = ScoutingReport
    form_class = ScoutingReportForm
    template_name = 'scouting/scoutingreport_form.html'
    success_message = 'Relatório de scouting atualizado com sucesso.'
    login_url = reverse_lazy('accounts:login')

    def get_success_url(self):
        return reverse('scouting:report_detail', args=[self.object.pk])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['score_fields'] = ScoutingReportForm.SCORE_FIELDS
        return context


class ScoutingReportDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a scouting report after confirmation."""

    model = ScoutingReport
    template_name = 'scouting/scoutingreport_confirm_delete.html'
    login_url = reverse_lazy('accounts:login')

    def get_success_url(self):
        player_id = self.object.player_id
        return reverse('scouting:player_detail', args=[player_id])

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Relatório de scouting removido com sucesso.')
        return super().delete(request, *args, **kwargs)


class ExportScoutingReportsView(LoginRequiredMixin, View):
    """Export scouting reports data to Excel format."""

    login_url = reverse_lazy('accounts:login')

    def get(self, request, *args, **kwargs):
        """Generate and return Excel file with scouting reports data."""
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Relatórios de Scouting'

        # Define header style
        header_fill = PatternFill(start_color='ec4899', end_color='ec4899', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Define headers (Portuguese as per UI convention)
        headers = [
            'ID',
            'Jogador',
            'Clube Atual',
            'Posição',
            'Status',
            'Data do Relatório',
            'Partida/Evento',
            'Avaliação Técnica',
            'Avaliação Física',
            'Avaliação Tática',
            'Avaliação Mental',
            'Potencial',
            'Nota Geral',
            'Pontos Fortes',
            'Pontos a Desenvolver',
            'Recomendação',
            'Observador',
            'Criado em',
            'Atualizado em',
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Fetch scouting reports data
        reports = (
            ScoutingReport.objects
            .select_related('player', 'created_by')
            .order_by('-report_date', 'player__name')
        )

        # Write data rows
        for row_num, report in enumerate(reports, 2):
            ws.cell(row=row_num, column=1, value=report.id)
            ws.cell(row=row_num, column=2, value=report.player.name)
            ws.cell(row=row_num, column=3, value=report.player.current_club)
            ws.cell(row=row_num, column=4, value=report.player.get_position_display())
            ws.cell(row=row_num, column=5, value=report.player.get_status_display())
            ws.cell(row=row_num, column=6, value=report.report_date.strftime('%d/%m/%Y'))
            ws.cell(row=row_num, column=7, value=report.match_or_event)
            ws.cell(row=row_num, column=8, value=report.technical_score)
            ws.cell(row=row_num, column=9, value=report.physical_score)
            ws.cell(row=row_num, column=10, value=report.tactical_score)
            ws.cell(row=row_num, column=11, value=report.mental_score)
            ws.cell(row=row_num, column=12, value=report.potential_score)
            ws.cell(row=row_num, column=13, value=round(report.overall_score(), 2))
            ws.cell(row=row_num, column=14, value=report.strengths)
            ws.cell(row=row_num, column=15, value=report.weaknesses)
            ws.cell(row=row_num, column=16, value=report.recommendation)
            ws.cell(row=row_num, column=17, value=report.created_by.email)
            ws.cell(row=row_num, column=18, value=report.created_at.strftime('%d/%m/%Y %H:%M:%S'))
            ws.cell(row=row_num, column=19, value=report.updated_at.strftime('%d/%m/%Y %H:%M:%S'))

        # Adjust column widths
        column_widths = [8, 30, 25, 20, 18, 18, 35, 18, 18, 18, 18, 15, 15, 40, 40, 40, 30, 22, 22]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        filename = f'relatorios_scouting_{timestamp}.xlsx'

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'

        # Save workbook to response
        wb.save(response)

        return response
