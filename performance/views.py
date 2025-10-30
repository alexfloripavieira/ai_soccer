from collections import defaultdict
from datetime import date, datetime, timedelta

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.db.models import Avg, Count, Max, Sum
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import CreateView, DeleteView, DetailView, ListView, TemplateView, UpdateView

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ml_models.injury_predictor import (
    MINIMUM_SAMPLES,
    evaluate_injury_risk_for_team,
    get_injury_predictor,
)

from performance.forms import AthleteForm, InjuryRecordForm, TrainingLoadForm
from performance.models import Athlete, InjuryRecord, TrainingLoad


class AthleteListView(LoginRequiredMixin, ListView):
    """List all athletes with search and filter capabilities."""

    model = Athlete
    template_name = 'performance/athlete_list.html'
    context_object_name = 'athletes'
    paginate_by = 20
    login_url = '/login/'

    def get_queryset(self):
        """
        Return filtered queryset based on search query and position filter.

        Supports:
        - 'q' parameter: Search by athlete name (case-insensitive)
        - 'position' parameter: Filter by position code (GK, DF, MF, FW)
        """
        queryset = super().get_queryset()

        # Search by name
        search_query = self.request.GET.get('q', '').strip()
        if search_query:
            queryset = queryset.filter(name__icontains=search_query)

        # Filter by position
        position_filter = self.request.GET.get('position', '').strip()
        if position_filter:
            queryset = queryset.filter(position=position_filter)

        return queryset

    def get_context_data(self, **kwargs):
        """Add position choices and current filters to context."""
        context = super().get_context_data(**kwargs)

        # Add position choices for filter dropdown
        context['position_choices'] = Athlete.POSITION_CHOICES

        # Add current filter values to preserve them in the template
        context['current_search'] = self.request.GET.get('q', '')
        context['current_position'] = self.request.GET.get('position', '')

        return context


class AthleteCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    """Create new athletes and bind them to the current user."""

    model = Athlete
    form_class = AthleteForm
    template_name = 'performance/athlete_form.html'
    success_url = reverse_lazy('performance:athlete_list')
    success_message = 'Atleta %(name)s cadastrado com sucesso!'
    login_url = reverse_lazy('accounts:login')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class AthleteDetailView(LoginRequiredMixin, DetailView):
    """Display detailed information about a single athlete."""

    model = Athlete
    template_name = 'performance/athlete_detail.html'
    context_object_name = 'athlete'
    login_url = reverse_lazy('accounts:login')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        latest_training_loads = list(
            self.object.training_loads.select_related('athlete')
            .order_by('-training_date')[:10]
        )
        aggregates = self.object.training_loads.aggregate(
            avg_duration=Avg('duration_minutes'),
            total_distance=Sum('distance_km'),
        )
        intensity_map = {
            TrainingLoad.INTENSITY_LOW: 1,
            TrainingLoad.INTENSITY_MEDIUM: 2,
            TrainingLoad.INTENSITY_HIGH: 3,
            TrainingLoad.INTENSITY_VERY_HIGH: 4,
        }

        injury_records = list(
            self.object.injury_records.select_related('created_by')
            .order_by('-injury_date', '-created_at')
        )
        active_injuries = [record for record in injury_records if record.actual_return is None]
        total_days_out = sum(record.days_out() for record in injury_records)
        injury_tab = self.request.GET.get('tab', 'training')
        if injury_tab not in {'training', 'injuries'}:
            injury_tab = 'training'

        context.update(
            {
                'latest_training_loads': latest_training_loads,
                'has_training_loads': bool(latest_training_loads),
                'training_load_average_duration': (
                    float(aggregates['avg_duration']) if aggregates['avg_duration'] else 0.0
                ),
                'training_load_total_distance': (
                    float(aggregates['total_distance']) if aggregates['total_distance'] else 0.0
                ),
                'training_load_count': self.object.training_loads.count(),
                'training_load_list_url': reverse_lazy('performance:training_load_list'),
                'training_load_chart_data': [
                    {
                        'date': load.training_date.strftime('%d/%m'),
                        'intensity': intensity_map.get(load.intensity_level, 0),
                        'label': load.get_intensity_level_display(),
                    }
                    for load in latest_training_loads
                ],
                'injury_records': injury_records,
                'has_injury_records': bool(injury_records),
                'is_currently_injured': bool(active_injuries),
                'current_injury': active_injuries[0] if active_injuries else None,
                'active_injuries': active_injuries,
                'injury_total_days_out': total_days_out,
                'injury_total_count': len(injury_records),
                'injury_list_url': reverse_lazy('performance:injury_record_list'),
                'injury_tab': injury_tab,
            }
        )
        return context


class AthleteUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    """Allow editing of existing athlete records."""

    model = Athlete
    form_class = AthleteForm
    template_name = 'performance/athlete_form.html'
    success_message = 'Atleta %(name)s atualizado com sucesso!'
    login_url = reverse_lazy('accounts:login')

    def get_success_url(self):
        return reverse_lazy('performance:athlete_detail', args=[self.object.pk])


class AthleteDeleteView(LoginRequiredMixin, DeleteView):
    """Handle confirmation and removal of an athlete."""

    model = Athlete
    template_name = 'performance/athlete_confirm_delete.html'
    login_url = reverse_lazy('accounts:login')
    success_url = reverse_lazy('performance:athlete_list')

    def delete(self, request, *args, **kwargs):
        athlete_name = self.get_object().name
        messages.success(request, f'Atleta {athlete_name} excluído com sucesso!')
        return super().delete(request, *args, **kwargs)


class TrainingLoadListView(LoginRequiredMixin, ListView):
    """List training loads with filters for athlete and period."""

    model = TrainingLoad
    template_name = 'performance/training_load_list.html'
    context_object_name = 'training_loads'
    paginate_by = 15
    login_url = reverse_lazy('accounts:login')

    def get_queryset(self):
        queryset = (
            super()
            .get_queryset()
            .select_related('athlete', 'created_by')
            .order_by('-training_date', '-created_at')
        )
        athlete_id = self.request.GET.get('athlete')
        if athlete_id:
            queryset = queryset.filter(athlete_id=athlete_id)

        period = self.request.GET.get('period')
        if period:
            today = date.today()
            periods = {
                'last_7_days': today - timedelta(days=7),
                'last_30_days': today - timedelta(days=30),
                'last_90_days': today - timedelta(days=90),
            }
            start_date = periods.get(period)
            if start_date:
                queryset = queryset.filter(training_date__gte=start_date)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['athletes'] = Athlete.objects.order_by('name')
        context['selected_athlete'] = self.request.GET.get('athlete', '')
        context['selected_period'] = self.request.GET.get('period', '')
        context['period_options'] = TrainingLoadForm.PERIOD_CHOICES
        context['params'] = self.request.GET.copy()
        context['params'].pop('page', None)
        context['total_records'] = self.get_queryset().count()
        aggregates = self.get_queryset().aggregate(
            avg_duration=Avg('duration_minutes'),
            latest_training_date=Max('training_date'),
        )
        context['average_duration'] = (
            float(aggregates['avg_duration']) if aggregates['avg_duration'] else 0.0
        )
        context['latest_training_date'] = aggregates['latest_training_date']
        return context


class TrainingLoadCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    """Create a new training load entry."""

    model = TrainingLoad
    form_class = TrainingLoadForm
    template_name = 'performance/training_load_form.html'
    login_url = reverse_lazy('accounts:login')
    success_url = reverse_lazy('performance:training_load_list')
    success_message = 'Carga de treino registrada com sucesso.'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def get_form(self, *args, **kwargs):
        form = super().get_form(*args, **kwargs)
        form.fields['athlete'].queryset = Athlete.objects.order_by('name')
        return form


class TrainingLoadUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    """Update an existing training load."""

    model = TrainingLoad
    form_class = TrainingLoadForm
    template_name = 'performance/training_load_form.html'
    login_url = reverse_lazy('accounts:login')
    success_url = reverse_lazy('performance:training_load_list')
    success_message = 'Carga de treino atualizada com sucesso.'

    def get_form(self, *args, **kwargs):
        form = super().get_form(*args, **kwargs)
        form.fields['athlete'].queryset = Athlete.objects.order_by('name')
        return form


class TrainingLoadDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a training load after confirmation."""

    model = TrainingLoad
    template_name = 'performance/training_load_confirm_delete.html'
    login_url = reverse_lazy('accounts:login')
    success_url = reverse_lazy('performance:training_load_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Carga de treino removida com sucesso.')
        return super().delete(request, *args, **kwargs)


class InjuryRecordListView(LoginRequiredMixin, ListView):
    """Display registered injuries with filtering options."""

    model = InjuryRecord
    template_name = 'performance/injury_record_list.html'
    context_object_name = 'injury_records'
    paginate_by = 15
    login_url = reverse_lazy('accounts:login')

    def get_queryset(self):
        queryset = (
            super()
            .get_queryset()
            .select_related('athlete', 'created_by')
            .order_by('-injury_date', '-created_at')
        )
        athlete_id = self.request.GET.get('athlete')
        severity = self.request.GET.get('severity')

        if athlete_id:
            queryset = queryset.filter(athlete_id=athlete_id)

        if severity:
            queryset = queryset.filter(severity_level=severity)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        filtered_queryset = self.get_queryset()
        context.update(
            {
                'athletes': Athlete.objects.order_by('name'),
                'severity_choices': InjuryRecord.SEVERITY_LEVEL_CHOICES,
                'selected_athlete': self.request.GET.get('athlete', ''),
                'selected_severity': self.request.GET.get('severity', ''),
                'total_records': filtered_queryset.count(),
                'active_cases': filtered_queryset.filter(actual_return__isnull=True).count(),
                'recent_injury': filtered_queryset.first(),
                'params': self.request.GET.copy(),
            }
        )
        context['params'].pop('page', None)
        return context


class InjuryRecordCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    """Register a new injury record."""

    model = InjuryRecord
    form_class = InjuryRecordForm
    template_name = 'performance/injury_record_form.html'
    login_url = reverse_lazy('accounts:login')
    success_url = reverse_lazy('performance:injury_record_list')
    success_message = 'Registro de lesão criado com sucesso.'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def get_form(self, *args, **kwargs):
        form = super().get_form(*args, **kwargs)
        form.fields['athlete'].queryset = Athlete.objects.order_by('name')
        return form


class InjuryRecordUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    """Update an existing injury record."""

    model = InjuryRecord
    form_class = InjuryRecordForm
    template_name = 'performance/injury_record_form.html'
    login_url = reverse_lazy('accounts:login')
    success_url = reverse_lazy('performance:injury_record_list')
    success_message = 'Registro de lesão atualizado com sucesso.'

    def get_form(self, *args, **kwargs):
        form = super().get_form(*args, **kwargs)
        form.fields['athlete'].queryset = Athlete.objects.order_by('name')
        return form


class InjuryRecordDeleteView(LoginRequiredMixin, DeleteView):
    """Remove an injury record after confirmation."""

    model = InjuryRecord
    template_name = 'performance/injury_record_confirm_delete.html'
    login_url = reverse_lazy('accounts:login')
    success_url = reverse_lazy('performance:injury_record_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Registro de lesão removido com sucesso.')
        return super().delete(request, *args, **kwargs)


class PerformanceDashboardView(LoginRequiredMixin, TemplateView):
    """Provide consolidated performance insights for coaching staff."""

    template_name = 'performance/performance_dashboard.html'
    login_url = reverse_lazy('accounts:login')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        total_athletes = Athlete.objects.count()
        injured_athlete_count = (
            InjuryRecord.objects.filter(actual_return__isnull=True)
            .values('athlete_id')
            .distinct()
            .count()
        )

        athlete_birth_dates = Athlete.objects.only('birth_date')
        ages = [athlete.age() for athlete in athlete_birth_dates]
        average_age = round(sum(ages) / len(ages), 1) if ages else None

        latest_training_loads = list(
            TrainingLoad.objects.select_related('athlete', 'created_by')
            .order_by('-training_date', '-created_at')[:5]
        )
        recent_injuries = list(
            InjuryRecord.objects.select_related('athlete', 'created_by')
            .order_by('-injury_date', '-created_at')[:5]
        )

        one_week_ago = date.today() - timedelta(days=7)
        recent_loads = (
            TrainingLoad.objects.filter(training_date__gte=one_week_ago)
            .select_related('athlete')
            .order_by('-training_date')
        )
        alert_map = defaultdict(lambda: {
            'athlete': None,
            'athlete_id': None,
            'sessions': 0,
            'high_sessions': 0,
            'total_duration': 0,
            'total_distance': 0.0,
            'fatigue_sum': 0.0,
            'last_training_date': None,
        })
        for load in recent_loads:
            entry = alert_map[load.athlete_id]
            entry['athlete'] = load.athlete
            entry['athlete_id'] = load.athlete_id
            entry['sessions'] += 1
            if load.intensity_level in [TrainingLoad.INTENSITY_HIGH, TrainingLoad.INTENSITY_VERY_HIGH]:
                entry['high_sessions'] += 1
            entry['total_duration'] += load.duration_minutes
            entry['total_distance'] += float(load.distance_km)
            entry['fatigue_sum'] += load.fatigue_index()
            if entry['last_training_date'] is None or load.training_date > entry['last_training_date']:
                entry['last_training_date'] = load.training_date

        alerts = []
        for data in alert_map.values():
            if data['high_sessions'] >= 2 or data['fatigue_sum'] >= 6 or data['total_duration'] >= 360:
                alerts.append(data)

        alerts.sort(key=lambda item: (item['high_sessions'], item['fatigue_sum'], item['total_duration']), reverse=True)
        alerts = alerts[:5]

        activity_items = []
        for load in latest_training_loads:
            activity_items.append(
                {
                    'date': load.training_date,
                    'type': 'training',
                    'label': 'Carga de treino',
                    'athlete': load.athlete,
                    'intensity': load.get_intensity_level_display(),
                    'description': f'{load.duration_minutes} min • {load.distance_km} km',
                }
            )
        for injury in recent_injuries:
            status_label = 'Recuperado' if injury.actual_return else 'Em recuperação'
            activity_items.append(
                {
                    'date': injury.injury_date,
                    'type': 'injury',
                    'label': 'Registro de lesão',
                    'athlete': injury.athlete,
                    'intensity': status_label,
                    'description': f'{injury.get_injury_type_display()} • {injury.get_body_part_display()}',
                }
            )
        activity_items.sort(key=lambda item: item['date'], reverse=True)
        latest_activities = activity_items[:8]

        context.update(
            {
                'total_athletes': total_athletes,
                'injured_athletes': injured_athlete_count,
                'active_athletes': max(total_athletes - injured_athlete_count, 0),
                'average_age': average_age,
                'latest_training_loads': latest_training_loads,
                'recent_injuries': recent_injuries,
                'alerts': alerts,
                'latest_activities': latest_activities,
                'has_data': bool(latest_training_loads or recent_injuries or alerts),
            }
        )
        return context


class InjuryRiskView(LoginRequiredMixin, TemplateView):
    """Display injury risk predictions generated by the ML model."""

    template_name = 'performance/injury_risk.html'
    login_url = reverse_lazy('accounts:login')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        predictor = get_injury_predictor()
        predictions = evaluate_injury_risk_for_team()
        labels_in_order = ['Baixo risco', 'Risco moderado', 'Alto risco']
        risk_distribution = {label: 0 for label in labels_in_order}
        for result in predictions:
            risk_distribution[result.risk_label] = risk_distribution.get(result.risk_label, 0) + 1
        distribution_entries = []
        total = len(predictions)
        for label in labels_in_order:
            count = risk_distribution[label]
            percentage = (count / total * 100) if total else 0
            distribution_entries.append(
                {
                    'label': label,
                    'count': count,
                    'percentage': percentage,
                }
            )

        highest_risk_chart = None
        if predictions:
            top_entry = predictions[0]
            highest_risk_chart = {
                'labels': [
                    'Duração média (min)',
                    'Distância média (km)',
                    'Sessões (28 dias)',
                    'Índice A/C',
                ],
                'values': [
                    min(top_entry.average_duration / 120, 1) * 100 if top_entry.average_duration else 0,
                    min(top_entry.average_distance / 12, 1) * 100 if top_entry.average_distance else 0,
                    min(top_entry.session_frequency / 16, 1) * 100 if top_entry.session_frequency else 0,
                    min(top_entry.acute_chronic_ratio / 2.5, 1) * 100 if top_entry.acute_chronic_ratio else 0,
                ],
            }

        context.update(
            {
                'predictions': predictions,
                'has_predictions': bool(predictions),
                'risk_distribution_entries': distribution_entries,
                'top_risks': predictions[:3],
                'highest_risk': predictions[0] if predictions else None,
                'highest_risk_chart': highest_risk_chart,
                'total_evaluated': len(predictions),
                'trained_samples': predictor.trained_samples,
                'validation_percentage': (
                    predictor.validation_accuracy * 100 if predictor.validation_accuracy is not None else None
                ),
                'minimum_samples': MINIMUM_SAMPLES,
                'insufficient_training_data': predictor.trained_samples < MINIMUM_SAMPLES,
            }
        )
        return context


class AthleteValuationView(LoginRequiredMixin, ListView):
    """Display athlete valuation data with market value aggregations."""

    model = Athlete
    template_name = 'performance/athlete_valuation.html'
    context_object_name = 'athletes'
    login_url = reverse_lazy('accounts:login')

    def get_queryset(self):
        """Return all athletes ordered by market value (highest first)."""
        return Athlete.objects.order_by('-market_value', 'name')

    def get_context_data(self, **kwargs):
        """Add valuation aggregations and statistics to context."""
        context = super().get_context_data(**kwargs)

        # Aggregate calculations
        aggregates = Athlete.objects.aggregate(
            total_squad_value=Sum('market_value'),
            average_value=Avg('market_value'),
        )

        # Athletes with and without market value
        athletes_with_value = Athlete.objects.filter(market_value__isnull=False)
        athletes_without_value_count = Athlete.objects.filter(market_value__isnull=True).count()

        context.update(
            {
                'total_squad_value': aggregates['total_squad_value'] or 0,
                'average_value': aggregates['average_value'] or 0,
                'athletes_with_value': athletes_with_value,
                'athletes_without_value': athletes_without_value_count,
            }
        )

        return context


class ExportAthletesView(LoginRequiredMixin, View):
    """Export athletes data to Excel format."""

    login_url = reverse_lazy('accounts:login')

    def get(self, request, *args, **kwargs):
        """Generate and return Excel file with athletes data."""
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Atletas'

        # Define header style
        header_fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Define headers (Portuguese as per UI convention)
        headers = [
            'ID',
            'Nome',
            'Data de Nascimento',
            'Idade',
            'Posição',
            'Nacionalidade',
            'Altura (cm)',
            'Peso (kg)',
            'Valor de Mercado (R$)',
            'Criado por',
            'Criado em',
            'Atualizado em',
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Fetch athletes data
        athletes = Athlete.objects.select_related('created_by').order_by('name')

        # Write data rows
        for row_num, athlete in enumerate(athletes, 2):
            ws.cell(row=row_num, column=1, value=athlete.id)
            ws.cell(row=row_num, column=2, value=athlete.name)
            ws.cell(row=row_num, column=3, value=athlete.birth_date.strftime('%d/%m/%Y'))
            ws.cell(row=row_num, column=4, value=athlete.age())
            ws.cell(row=row_num, column=5, value=athlete.get_position_display())
            ws.cell(row=row_num, column=6, value=athlete.nationality)
            ws.cell(row=row_num, column=7, value=athlete.height)
            ws.cell(row=row_num, column=8, value=athlete.weight)
            ws.cell(row=row_num, column=9, value=float(athlete.market_value) if athlete.market_value else '')
            ws.cell(row=row_num, column=10, value=athlete.created_by.email)
            ws.cell(row=row_num, column=11, value=athlete.created_at.strftime('%d/%m/%Y %H:%M:%S'))
            ws.cell(row=row_num, column=12, value=athlete.updated_at.strftime('%d/%m/%Y %H:%M:%S'))

        # Adjust column widths
        column_widths = [8, 30, 20, 10, 20, 20, 15, 15, 25, 30, 22, 22]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        filename = f'atletas_{timestamp}.xlsx'

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'

        # Save workbook to response
        wb.save(response)

        return response


class ExportTrainingLoadsView(LoginRequiredMixin, View):
    """Export training loads data to Excel format."""

    login_url = reverse_lazy('accounts:login')

    def get(self, request, *args, **kwargs):
        """Generate and return Excel file with training loads data."""
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Cargas de Treino'

        # Define header style
        header_fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Define headers (Portuguese as per UI convention)
        headers = [
            'ID',
            'Atleta',
            'Data do Treino',
            'Duração (min)',
            'Distância (km)',
            'FC Média',
            'FC Máxima',
            'Intensidade',
            'Índice de Fadiga',
            'Criado por',
            'Criado em',
            'Atualizado em',
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Fetch training loads data
        training_loads = (
            TrainingLoad.objects
            .select_related('athlete', 'created_by')
            .order_by('-training_date', 'athlete__name')
        )

        # Write data rows
        for row_num, load in enumerate(training_loads, 2):
            ws.cell(row=row_num, column=1, value=load.id)
            ws.cell(row=row_num, column=2, value=load.athlete.name)
            ws.cell(row=row_num, column=3, value=load.training_date.strftime('%d/%m/%Y'))
            ws.cell(row=row_num, column=4, value=load.duration_minutes)
            ws.cell(row=row_num, column=5, value=float(load.distance_km))
            ws.cell(row=row_num, column=6, value=load.heart_rate_avg if load.heart_rate_avg else '')
            ws.cell(row=row_num, column=7, value=load.heart_rate_max if load.heart_rate_max else '')
            ws.cell(row=row_num, column=8, value=load.get_intensity_level_display())
            ws.cell(row=row_num, column=9, value=round(load.fatigue_index(), 2))
            ws.cell(row=row_num, column=10, value=load.created_by.email)
            ws.cell(row=row_num, column=11, value=load.created_at.strftime('%d/%m/%Y %H:%M:%S'))
            ws.cell(row=row_num, column=12, value=load.updated_at.strftime('%d/%m/%Y %H:%M:%S'))

        # Adjust column widths
        column_widths = [8, 30, 18, 15, 18, 12, 12, 18, 20, 30, 22, 22]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        filename = f'cargas_treino_{timestamp}.xlsx'

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'

        # Save workbook to response
        wb.save(response)

        return response
